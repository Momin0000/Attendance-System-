"""
Phase 3 - Reports Router
Add to main.py: from app.reports.router import router as reports_router
                app.include_router(reports_router)
"""
import csv
import io
import os
from typing import Optional
from datetime import date
from pathlib import Path

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from app.database.session import get_db
from app.models.models import Student, Attendance, Campus
from app.core.security import get_current_user, get_operator_or_above

router = APIRouter(prefix="/reports", tags=["Reports"])

HEADER_COLOR = "003366"
HEADER_FONT_COLOR = "FFFFFF"


def _query_attendance(
    db: Session,
    campus_id: Optional[int],
    batch_id: Optional[int],
    course_id: Optional[int],
    from_date: Optional[date],
    to_date: Optional[date],
):
    q = db.query(Attendance)
    if campus_id:
        q = q.filter(Attendance.campus_id == campus_id)
    if batch_id:
        q = q.filter(Attendance.batch_id == batch_id)
    if from_date:
        q = q.filter(Attendance.attendance_date >= from_date)
    if to_date:
        q = q.filter(Attendance.attendance_date <= to_date)
    return q.order_by(Attendance.attendance_date.desc()).all()


def _student_map(db: Session):
    students = db.query(Student).all()
    return {s.id: s for s in students}


@router.get("/attendance/excel")
async def attendance_excel(
    campus_id: Optional[int] = Query(None),
    batch_id: Optional[int] = Query(None),
    course_id: Optional[int] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_operator_or_above),
):
    records = _query_attendance(db, campus_id, batch_id, course_id, from_date, to_date)
    smap = _student_map(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(color=HEADER_FONT_COLOR, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws.merge_cells("A1:H1")
    ws["A1"].value = "BANO QABIL PAKISTAN - ATTENDANCE REPORT"
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_COLOR)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ["#", "Student ID", "Full Name", "Email", "Date", "Check-In", "Status", "Method"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[2].height = 20

    for i, r in enumerate(records, 1):
        s = smap.get(r.student_db_id)
        row_data = [
            i, r.student_id_code,
            s.full_name if s else "—", s.email if s else "—",
            str(r.attendance_date),
            str(r.check_in_time)[:19] if r.check_in_time else "—",
            r.status.upper(), r.scan_method.upper() if hasattr(r, "scan_method") else "QR",
        ]
        row_num = i + 2
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if r.status == "present":
                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            elif r.status == "absent":
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    for col, width in zip(range(1, 9), [5, 18, 28, 32, 12, 22, 10, 10]):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"},
    )


@router.get("/attendance/csv")
async def attendance_csv(
    campus_id: Optional[int] = Query(None),
    batch_id: Optional[int] = Query(None),
    course_id: Optional[int] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_operator_or_above),
):
    records = _query_attendance(db, campus_id, batch_id, course_id, from_date, to_date)
    smap = _student_map(db)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["#", "Student ID", "Full Name", "Email", "Date", "Check-In", "Status"])
    for i, r in enumerate(records, 1):
        s = smap.get(r.student_db_id)
        writer.writerow([
            i, r.student_id_code,
            s.full_name if s else "—", s.email if s else "—",
            str(r.attendance_date),
            str(r.check_in_time)[:19] if r.check_in_time else "—",
            r.status,
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance_report.csv"},
    )


@router.get("/attendance/pdf")
async def attendance_pdf(
    campus_id: Optional[int] = Query(None),
    batch_id: Optional[int] = Query(None),
    course_id: Optional[int] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_operator_or_above),
):
    records = _query_attendance(db, campus_id, batch_id, course_id, from_date, to_date)
    smap = _student_map(db)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("BANO QABIL PAKISTAN", ParagraphStyle("t", parent=styles["Title"], fontSize=14,
                  textColor=colors.HexColor("#003366"), alignment=TA_CENTER)),
        Paragraph("Attendance Report", ParagraphStyle("s", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER)),
        Spacer(1, 6*mm),
    ]

    table_data = [["#", "Student ID", "Name", "Email", "Date", "Status"]]
    for i, r in enumerate(records, 1):
        s = smap.get(r.student_db_id)
        table_data.append([str(i), r.student_id_code,
                           s.full_name if s else "—", s.email if s else "—",
                           str(r.attendance_date), r.status.upper()])

    t = Table(table_data, colWidths=[10*mm, 32*mm, 55*mm, 65*mm, 28*mm, 22*mm], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=attendance_report.pdf"},
    )


@router.get("/students/excel")
async def students_excel(
    campus_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_operator_or_above),
):
    q = db.query(Student).filter(Student.is_deleted == False)
    if campus_id:
        q = q.filter(Student.campus_id == campus_id)
    students = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(color=HEADER_FONT_COLOR, bold=True)
    headers = ["#", "Student ID", "Full Name", "Father Name", "CNIC", "Phone", "Email", "Gender", "Enrollment Date", "Active"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, s in enumerate(students, 1):
        ws.append([
            i, s.student_id, s.full_name, s.father_name or "—",
            s.cnic or "—", s.phone or "—", s.email or "—",
            s.gender or "—", str(s.enrollment_date), "Yes" if s.is_active else "No",
        ])

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_report.xlsx"},
    )
