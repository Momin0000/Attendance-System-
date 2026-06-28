import os
import io
import csv
from uuid import UUID
from datetime import date
from typing import Optional
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from app.models.student import Student
from app.models.attendance import AttendanceRecord


REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)


def _get_attendance_data(
    db: Session,
    batch_id: Optional[UUID] = None,
    campus_id: Optional[UUID] = None,
    course_id: Optional[UUID] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
):
    query = db.query(AttendanceRecord)
    if batch_id:
        query = query.filter(AttendanceRecord.batch_id == batch_id)
    if campus_id:
        query = query.filter(AttendanceRecord.campus_id == campus_id)
    if course_id:
        query = query.filter(AttendanceRecord.course_id == course_id)
    if from_date:
        query = query.filter(AttendanceRecord.attendance_date >= from_date)
    if to_date:
        query = query.filter(AttendanceRecord.attendance_date <= to_date)
    return query.order_by(AttendanceRecord.attendance_date.desc()).all()


def export_attendance_excel(
    db: Session,
    batch_id: Optional[UUID] = None,
    campus_id: Optional[UUID] = None,
    course_id: Optional[UUID] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
) -> str:
    records = _get_attendance_data(db, batch_id, campus_id, course_id, from_date, to_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header styling
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "BANO QABIL PAKISTAN - ATTENDANCE REPORT"
    title_cell.font = Font(bold=True, size=14, color="003366")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 25

    headers = ["#", "Student ID", "Student Name", "Email", "Date", "Check-In Time", "Status", "Method"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[2].height = 20

    for row_idx, record in enumerate(records, start=3):
        student = db.query(Student).filter(Student.id == record.student_id).first()
        full_name = f"{student.first_name} {student.last_name}" if student else "N/A"
        row_data = [
            row_idx - 2,
            student.student_id if student else "N/A",
            full_name,
            student.email if student else "N/A",
            str(record.attendance_date),
            str(record.check_in_time) if record.check_in_time else "N/A",
            record.status.upper(),
            record.scan_method.upper(),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if record.status == "present":
                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            elif record.status == "absent":
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    # Auto column width
    col_widths = [5, 18, 25, 35, 12, 22, 10, 10]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    filepath = os.path.join(REPORTS_DIR, "attendance_report.xlsx")
    wb.save(filepath)
    return filepath


def export_attendance_csv(
    db: Session,
    batch_id: Optional[UUID] = None,
    campus_id: Optional[UUID] = None,
    course_id: Optional[UUID] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
) -> str:
    records = _get_attendance_data(db, batch_id, campus_id, course_id, from_date, to_date)

    filepath = os.path.join(REPORTS_DIR, "attendance_report.csv")
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["#", "Student ID", "Student Name", "Email", "Date", "Check-In Time", "Status", "Method"])
        for idx, record in enumerate(records, start=1):
            student = db.query(Student).filter(Student.id == record.student_id).first()
            full_name = f"{student.first_name} {student.last_name}" if student else "N/A"
            writer.writerow([
                idx,
                student.student_id if student else "N/A",
                full_name,
                student.email if student else "N/A",
                str(record.attendance_date),
                str(record.check_in_time) if record.check_in_time else "N/A",
                record.status,
                record.scan_method,
            ])

    return filepath


def export_attendance_pdf(
    db: Session,
    batch_id: Optional[UUID] = None,
    campus_id: Optional[UUID] = None,
    course_id: Optional[UUID] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
) -> str:
    records = _get_attendance_data(db, batch_id, campus_id, course_id, from_date, to_date)

    filepath = os.path.join(REPORTS_DIR, "attendance_report.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=14, textColor=colors.HexColor("#003366"), alignment=TA_CENTER)

    elements = []
    elements.append(Paragraph("BANO QABIL PAKISTAN", title_style))
    elements.append(Paragraph("Attendance Report", ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER)))
    elements.append(Spacer(1, 5*mm))

    table_data = [["#", "Student ID", "Name", "Email", "Date", "Status", "Method"]]
    for idx, record in enumerate(records, start=1):
        student = db.query(Student).filter(Student.id == record.student_id).first()
        full_name = f"{student.first_name} {student.last_name}" if student else "N/A"
        table_data.append([
            str(idx),
            student.student_id if student else "N/A",
            full_name,
            student.email if student else "N/A",
            str(record.attendance_date),
            record.status.upper(),
            record.scan_method.upper(),
        ])

    col_widths = [10*mm, 30*mm, 45*mm, 60*mm, 25*mm, 20*mm, 20*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(t)
    doc.build(elements)
    return filepath


def export_students_excel(db: Session, campus_id: Optional[UUID] = None) -> str:
    query = db.query(Student)
    if campus_id:
        query = query.filter(Student.campus_id == campus_id)
    students = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["#", "Student ID", "First Name", "Last Name", "Email", "Phone", "CNIC", "Gender", "Enrollment Date", "Active"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=student.student_id)
        ws.cell(row=row, column=3, value=student.first_name)
        ws.cell(row=row, column=4, value=student.last_name)
        ws.cell(row=row, column=5, value=student.email)
        ws.cell(row=row, column=6, value=student.phone)
        ws.cell(row=row, column=7, value=student.cnic)
        ws.cell(row=row, column=8, value=student.gender)
        ws.cell(row=row, column=9, value=str(student.enrollment_date) if student.enrollment_date else "")
        ws.cell(row=row, column=10, value="Yes" if student.is_active else "No")

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18

    filepath = os.path.join(REPORTS_DIR, "students_report.xlsx")
    wb.save(filepath)
    return filepath
